import openpyxl
def getAllValues(path):
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    ar = []
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        ar.append(cell_obj.value)

    return ar
